import socket
import pandas as pd
import os
import time
from openpyxl import load_workbook
s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
s.bind(('', 54321))
name = ['Alice', 'Bob', 'Carol', 'Dave', 'Eve', 'Francis', 'Grace', 'Hans', 'Isabella']
while True:
    data, address = s.recvfrom(4096)
    print(address)
    text = data.decode()
    print(text)
    if text[0] == 'n':
        if not os.path.exists("./group.xlsx"):
            df = pd.DataFrame({'name' : [name[0]], 'ip' : [address[0]]})
            df.to_excel("./group.xlsx", sheet_name=f'{text[10:]}', index=False)
            print("success")
            continue
        book = load_workbook('./group.xlsx')
        group_writer = pd.ExcelWriter('./group.xlsx', engine='openpyxl')
        group_writer.book = book
        group_writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        if text[10:] in group_writer.sheets:
            print("should send back")
            group_writer.close()
            continue
        else:
            df = pd.DataFrame({'name' : [name[0]], 'ip' : [address[0]]})
            df.to_excel(group_writer, sheet_name=f'{text[10:]}', index=False)
            group_writer.save()
            print("success")
    elif text[0] == 'a':
        if not os.path.exists("./group.xlsx"):
            print("should send back")
            continue
        book = load_workbook('./group.xlsx')
        group_writer = pd.ExcelWriter('./group.xlsx', engine='openpyxl')
        group_writer.book = book
        group_writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        if not text[10:] in group_writer.sheets:
            print("should send back")
            group_writer.close()
            continue
        else:
            group_writer.close()
            df_old = pd.DataFrame(pd.read_excel('./group.xlsx', sheet_name=f'{text[10:]}'))
            find = False
            for i in range(df_old.shape[0]):
                tmp = df_old.iat[i, 1]
                if tmp == address[0]:
                    print("should send back")
                    find = True
                    break
            if find:
                continue
            n = len(df_old.index)
            print(n)
            print(name[n])
            df = pd.DataFrame({'name' : [name[n]], 'ip' : [address[0]]})
            book = load_workbook('./group.xlsx')
            group_writer = pd.ExcelWriter('./group.xlsx', engine='openpyxl')
            group_writer.book = book
            group_writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df_rows = df_old.shape[0]
            df.to_excel(group_writer, sheet_name=f'{text[10:]}', startrow=df_rows+1,header=False, index=False)
            group_writer.save()
            print("success")
    elif text[0] == 's':
        ls = text.split(' ')
        length = len(ls[1])
        if not os.path.exists("./group.xlsx"):
            print("should send back")
            continue
        book = load_workbook('./group.xlsx')
        group_writer = pd.ExcelWriter('./group.xlsx', engine='openpyxl')
        group_writer.book = book
        group_writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        if not text[5:5+length] in group_writer.sheets:
            print("should send back")
            group_writer.close()
            continue
        group_writer.close()
        
        t = time.localtime()
        df_old = pd.DataFrame(pd.read_excel('./group.xlsx', sheet_name=f'{text[5:5+length]}'))
        name_now = ''
        for i in range(df_old.shape[0]):
                tmp = df_old.iat[i, 1]
                if tmp == address[0]:
                    name_now = df_old.iat[i, 0]
        if not os.path.exists("./data.xlsx"):
            df = pd.DataFrame({'name' : [name_now], 'message' : [text[length + 6:]], 'time' : [f'{t.tm_mon}-{t.tm_mday} {t.tm_hour}:{t.tm_min}']})
            df.to_excel("./data.xlsx", sheet_name=f'{text[5:5+length]}', index=False)
            print("success")
            continue
        book = load_workbook('./data.xlsx')
        data_writer = pd.ExcelWriter('./data.xlsx', engine='openpyxl')
        data_writer.book = book
        data_writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        if not text[5:5+length] in data_writer.sheets:
            df = pd.DataFrame({'name' : [name_now], 'message' : [text[length + 6:]], 'time' : ['{:0>2d}-{:0>2d} {:0>2d}:{:0>2d}'.format(t.tm_mon, t.tm_mday, t.tm_hour, t.tm_min)]})
            df.to_excel(data_writer, sheet_name=f'{text[5:5+length]}', index=False)
            data_writer.close()
            print("success")
            continue
        else:
            data_writer.close()
            df = pd.DataFrame({'name' : [name_now], 'message' : [text[length + 6:]], 'time' : ['{:0>2d}-{:0>2d} {:0>2d}:{:0>2d}'.format(t.tm_mon, t.tm_mday, t.tm_hour, t.tm_min)]})
            df_old = pd.DataFrame(pd.read_excel('data.xlsx', sheet_name=f'{text[5:5+length]}'))
            book = load_workbook('./data.xlsx')
            data_writer = pd.ExcelWriter('./data.xlsx', engine='openpyxl')
            data_writer.book = book
            data_writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df_rows = df_old.shape[0]
            df.to_excel(data_writer, sheet_name=f'{text[5:5+length]}', startrow=df_rows+1,header=False, index=False)
            data_writer.save()
            print("success")
    elif text[0] == 'r':
        if not os.path.exists("./data.xlsx"):
            print("should send back")
            continue
        book = load_workbook('./data.xlsx')
        data_writer = pd.ExcelWriter('./data.xlsx', engine='openpyxl')
        data_writer.book = book
        data_writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        if not text[7:] in data_writer.sheets:
            print("should send back")
            data_writer.close()
            continue
        else:
            data_writer.close()
            df = pd.DataFrame(pd.read_excel('data.xlsx', sheet_name=f'{text[7:]}'))
            message = ''
            if df.shape[1] < 10:
                for i in range(df.shape[0]):
                    for j in range(df.shape[1]):
                        message += df.iat[i, j]
                        if j != df.shape[1] - 1:
                            message += ' '
                    message += '\n'
            else:
                for i in range(df.shape[0] - 1, df.shape[0] - 11, -1):
                    for j in range(df.shape[1]):
                        message += df.iat[i, j]
                        if j != df.shape[1] - 1:
                            message += ' '
                    message += '\n'
            print(message, end='')
            print(address[0])
            s.sendto(message.encode(), (address[0], 12345))






